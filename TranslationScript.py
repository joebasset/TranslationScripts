import json
import os
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
import pandas as pd
import re


def parse_text_file(filename, encoding="utf-8"):
    try:
        with open(filename, "r", encoding=encoding) as file:
            lines = file.readlines()
        return lines
    except Exception as e:
        print(f"Error reading file '{filename}': {str(e)}")
        return None


def flatten_json(y):
    out = {}

    def flatten(x, name=""):
        if type(x) is dict:
            for a in x:
                flatten(x[a], name + a + "_")
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + "_")
                i += 1
        else:
            out[name[:-1]] = x

    flatten(y)
    return out


def is_empty_line(line):
    return all(char in ",{}" for char in line.strip())


def is_last_line(line):
    pattern = r"^\},$"
    match = re.search(pattern, line)
    return bool(match)


root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(
    title="Select a file",
    filetypes=[
        ("Text files", "*.txt"),
    ],
)

if not file_path:
    print("No file selected. Exiting.")
else:
    _, file_extension = os.path.splitext(file_path)
    if file_extension == ".json":
        with open(file_path, "r", encoding="utf-8") as json_file:
            data = json.load(json_file)
    elif file_extension in [".js", ".ts", ".txt"]:
        data = {"file_content": parse_text_file(file_path)}
    else:
        print(f"Unsupported file type: {file_extension}. Exiting.")
        data = {}

    if data:
        workbook = Workbook()
        sheet = workbook.active

        if "file_content" in data:
            sheet.cell(row=1, column=1, value="Main Key / Figma Screens")
            sheet.cell(row=1, column=2, value="Key")
            sheet.cell(row=1, column=3, value="EN")
            sheet.cell(row=1, column=4, value="FR")
            sheet.cell(row=1, column=5, value="AR")

            row = 3

            for line in data["file_content"]:
                col = 1
                line = line.strip()
                if is_last_line(line.strip()):
                    row += 2
                    continue
                if line and not is_empty_line(line) and ":" in line:
                    key, value = line.split(":", 1)
                    if is_empty_line(value.strip()):
                        sheet.cell(row=row, column=col, value=key.strip())
                        row += 1
                        continue
                    col += 1
                    sheet.cell(row=row, column=col, value=key.strip())
                    sheet.cell(
                        row=row,
                        column=3,
                        value=value.strip()
                        .replace("'", "")
                        .replace('"', "")
                        .replace(",", ""),
                    )
                    sheet.cell(
                        row=row,
                        column=4,
                        value="=GOOGLETRANSLATE(C" + str(row) + ', "en","fr")',
                    )
                    sheet.cell(
                        row=row,
                        column=5,
                        value="=GOOGLETRANSLATE(C" + str(row) + ', "en","ar")',
                    )

                    row += 1

        else:
            flat_data = flatten_json(data)

            row = 1
            for key, value in flat_data.items():
                sheet.cell(row=row, column=1, value=key)
                sheet.cell(row=row, column=3, value=value)
                row += 1

        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
        )

        if output_path:
            workbook.save(output_path)
            print(f'Excel file "{output_path}" created successfully.')


root.destroy()
